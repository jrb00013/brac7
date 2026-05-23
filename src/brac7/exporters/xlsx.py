"""Excel bracket workbook export."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from brac7.models import Bracket

THIN = Side(style="thin", color="D9E1F2")
GRID = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
ALT_FILL = PatternFill("solid", fgColor="EEF2FA")
BYE_FILL = PatternFill("solid", fgColor="E8F5E9")


def _style_header(ws, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def export_xlsx(bracket: Bracket, path: Path) -> Path:
    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = "Info"
    hf = Font(bold=True, size=12)

    meta = [
        ("Title", bracket.title),
        ("Format", bracket.format_label),
        ("Seeding", bracket.options.seeding.value),
        ("Byes", "Yes" if bracket.options.supports_byes else "No"),
        ("Max participants", bracket.options.max_participants or "—"),
        ("Bracket size", bracket.size),
        ("Entrants", len(bracket.participants)),
        ("Match format", bracket.options.match_format.value),
    ]
    for row, (k, v) in enumerate(meta, start=1):
        ws_meta.cell(row, 1, k).font = hf
        ws_meta.cell(row, 2, v)
    _auto_width(ws_meta)

    ws_p = wb.create_sheet("Participants")
    ws_p.append(["Seed", "Name"])
    _style_header(ws_p, 2)
    for i, p in enumerate(bracket.participants, start=2):
        ws_p.cell(i, 1, p.seed).border = GRID
        ws_p.cell(i, 2, p.name).border = GRID
        if i % 2 == 0:
            ws_p.cell(i, 1).fill = ALT_FILL
            ws_p.cell(i, 2).fill = ALT_FILL
    _auto_width(ws_p)

    ws_m = wb.create_sheet("Matches")
    headers = ["Round", "Match ID", "Bracket", "Slot A", "Slot B", "Bye", "Label", "Feeds From"]
    ws_m.append(headers)
    _style_header(ws_m, len(headers))
    for i, rnd in enumerate(bracket.rounds):
        for m in rnd.matches:
            row = ws_m.max_row + 1
            vals = [
                rnd.name,
                m.id,
                m.bracket,
                m.display_slot("a", bracket.options),
                m.display_slot("b", bracket.options),
                "Yes" if m.is_bye else "",
                m.label,
                ", ".join(m.feeds_from),
            ]
            for col, val in enumerate(vals, 1):
                cell = ws_m.cell(row, col, val)
                cell.border = GRID
                if m.is_bye:
                    cell.fill = BYE_FILL
    _auto_width(ws_m)

    ws_v = wb.create_sheet("Bracket View")
    ws_v.append(["Round", "Match", "Competitor 1", "vs", "Competitor 2"])
    _style_header(ws_v, 5)
    col = 1
    for rnd in bracket.rounds:
        row = 2
        for m in rnd.matches:
            ws_v.cell(row, col, rnd.name).border = GRID
            ws_v.cell(row, col + 1, m.id).border = GRID
            ws_v.cell(row, col + 2, m.display_slot("a", bracket.options)).border = GRID
            ws_v.cell(row, col + 3, "vs").border = GRID
            ws_v.cell(row, col + 4, m.display_slot("b", bracket.options)).border = GRID
            for c in range(col, col + 5):
                ws_v.cell(row, c).alignment = Alignment(vertical="center")
            row += 2
        col += 6
    _auto_width(ws_v)

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
